import pandas as pd
import shutil
from tkinter import filedialog
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from pathlib import Path


def member_list_import():
    # чтение файла
    filename = filedialog.askopenfilename()
    directory = filename.rsplit('/', 1)[0]
    data = pd.read_excel(filename)

    # извленчение нужных данных (имена, номера квартир, площадь)
    data = data.drop(data[data['Unnamed: 2'].isna()].index)
    data = data.iloc[2:].reset_index()
    members = pd.DataFrame(
        {'name': data['Unnamed: 3'] + ' ' + data['Unnamed: 4'] + ' ' + data['Unnamed: 5'],
         'flat': data['Unnamed: 2'],
         'square': data['Unnamed: 7']})
    members = members.astype({'square': 'float',
                              'flat': 'str',
                              'name': 'str'})
    members.square = members.square.map(lambda x: round(x, 1))
    print(members.head(2).to_string(), '\n', members.tail(2).to_string(), '\n')
    return members, directory


def number_of_q():
    while True:
        try:
            val = int(input('Количество вопросов?'))
            if val > 0:
                print(val, ' вопросов', '\n')
                return val
            else:
                print('некорректный ввод')
        except ValueError:
            print('некорректный ввод')


def decoder(a):
    code = {'0': '001',
            '1': '010',
            '2': '100'}
    result = ''.join(map(lambda x: code[x], list(a)))
    return list(map(int, result))


def template_answer(n):
    while True:
        temp = input('Шаблон ответа?')
        if (len(temp) != n) | (temp.strip('012') != ''):
            print('некорректный ввод')
        else:
            return temp


def template(answer, columns):
    # запрос списка квартир для шаблона, формирование шаблона
    tmp_list = input('Список квартир(через запятую)?')
    result = pd.DataFrame(columns=columns)
    result.flat = list(tmp_list.split(','))
    result.iloc[:, 1:5] = [1, 1, 0, 0]
    result.iloc[:, 5:] = decoder(answer)
    return result


def template_check(flats, member_list, previous):
    # проверка шаблона на правильность

    # Проверка на то, имеются ли введенные квартиры в реестре
    if not (flats.flat.isin(member_list.flat).all()):
        print('в реестре отсутствуют квартиры:')
        for i in flats.loc[flats.flat.isin(member_list.flat) == False, 'flat'].values:
            print(i)
        print('они не будут учтены в голосовании:')
        flats_out = flats[flats.flat.isin(member_list.flat)]
        flats = flats_out

    # Проверка на то, не были ли введены данные квартиры на предыдущем этапе
    if flats.flat.isin(previous.flat).any():
        print('Эти квартиры были введены на предыдущем этапе:')
        for i in flats.loc[flats.flat.isin(previous.flat), 'flat'].values:
            print(i)
        print('они не будут учтены в голосовании:')
        flats_out = flats[flats.flat.isin(previous.flat) == False]
        flats = flats_out

    # Проверка на то, не введены ли квартиры дважды
    if flats.flat.duplicated().any():
        print('Номер квартиры введен дважды:')
        for i in flats.loc[flats.flat.duplicated(), 'flat'].values:
            print(i)
        print('Проверьте правильность ввода')
        flats = flats.drop_duplicates()

    # Подтверждение принятия шаблона
    i = 0
    while i == 0:
        confirm = input('Принять шаблон? (y/n)')
        if confirm == 'y':
            i = 1
        elif confirm == 'n':
            flats = flats.iloc[0:0]
            i = 1
        else:
            print('некорректный ввод')

    return flats


def template_enter(df, numquest):
    columns = ['flat', 'filled', '1', '2', '3']
    for i in range(numquest * 3):
        columns.append(str(i + 4))
    result = pd.DataFrame(columns=columns)

    i = 0
    while i == 0:
        answer = template_answer(numquest)
        new_result = template(answer, columns)
        new_result = template_check(new_result, df, result)
        result = pd.concat([result, new_result], axis=0)

        j = 0
        while j == 0:
            confirm = input('ввести еще один шаблон? (y/n)')
            if confirm == 'y':
                j = 1
            elif confirm == 'n':
                i, j = 1, 1
            else:
                print('некорректный ввод')
    df = df.merge(result, how='left', on='flat')
    print('переходим к построчному вводу')
    df = line_enter(df, numquest)
    return df


def single_line_print(s, answer):
    ans = pd.Series(list(answer))
    dec = pd.Series(['нет', 'возд', 'да'], index=['0', '1', '2'])
    col = ['собственник', '№ кв.']
    for i in range(len(ans)):
        col.append('Вопр.' + str(i + 1))
    ret = pd.DataFrame(columns=col, index=[1, ])
    ret.loc[1, ['собственник', '№ кв.']] = [s['name'], s['flat']]
    ret.iloc[0, 2:] = ans.map(dec).array
    print('\n', ret.to_string(index=False), '\n')
    return


def line_enter(df, num_question):
    idx_list = df[df['filled'].isna()].index

    i = 0
    while i < idx_list.size:
        answer = input(df.loc[idx_list[i], 'name'])
        if answer == 's':
            i = idx_list.size
            print('ввод прерван', '\n')
        elif answer == 'n':
            df.loc[idx_list[i], 'filled'] = 1
            df.loc[idx_list[i], ['1', '2', '3']] = [0, 0, 1]
            i += 1
            print('недействительный бюллетень', '\n')
        elif (answer == 'b') & (i == 0):
            pass
        elif (answer == 'b') & (i != 0):
            i -= 1
            df.iloc[idx_list[i], 7:] = None
            print('вернулись на предыдущий шаг', '\n')
        elif answer == '':
            df.loc[idx_list[i], 'filled'] = 1
            df.loc[idx_list[i], ['1', '2', '3']] = [0, 1, 0]
            i += 1
            print('не голосовал', '\n')
        elif (len(answer) != num_question) | (answer.strip('012') != ''):
            print('некорректный ввод', '\n')
        else:
            df.loc[idx_list[i], 'filled'] = 1
            df.loc[idx_list[i], ['1', '2', '3']] = [1, 0, 0]
            df.iloc[idx_list[i], 7:] = decoder(answer)
            single_line_print(df.loc[idx_list[i]], answer)
            i += 1
    return df


def single_line_enter(df, numquest):
    columns = ['filled', '1', '2', '3']
    for i in range(numquest * 3):
        columns.append(str(i + 4))
    df = pd.concat([df, pd.DataFrame(columns=columns)], axis=1)
    df = line_enter(df, numquest)
    return df


def result_analyse(df):
    total_square = df.square.sum()
    i = 0
    while i == 0:
        confirm = input('Ипользовать полученное значение для общей площади = '
                        '{} кв.м. (1) или стандартное значение 7779.1 кв.м. (2)?'
                        .format(str(round(total_square, 1))))
        if confirm == '1':
            i = 1
        elif confirm == '2':
            total_square = 7779.1
            i = 1
        else:
            print('Некорректный ввод. Введите 1 или 2.')
    print('\n', '\n')

    results_analyse = pd.DataFrame(index=[1, 2], columns=df.columns[3:])
    results_analyse.iloc[0] = df[df.columns[3:]].mul(df['square'], axis=0).sum()
    results_analyse.iloc[1, :3] = results_analyse.iloc[0, :3] / total_square * 100
    results_analyse.iloc[1, 3:] = results_analyse.iloc[0, 3:] / results_analyse.iloc[0, 0] * 100
    results_analyse.iloc[0] = results_analyse.iloc[0].apply(lambda x: round(x, 1))
    results_analyse.iloc[1] = results_analyse.iloc[1].apply(lambda x: str(round(x, 1)) + ' %')
    results_analyse['total'] = [round(total_square, 1), None]
    return results_analyse


def print_result(df1, df2):
    # Переименуем столбцы
    arrays = [['', ' ', 'площадь'],
              ['Собственник', '№ кв.', 'помещения']]
    col1 = pd.MultiIndex.from_arrays(arrays)
    questions = []
    for i in range(int((df1.shape[1] - 6) / 3)):
        questions.append('Вопрос ' + str(i + 1))
    col2 = pd.MultiIndex.from_product([['участие'], ['да', 'нет', 'недейств']])
    col3 = pd.MultiIndex.from_product([questions, ['Да', 'Возд', 'Нет']])
    col4 = pd.MultiIndex.from_arrays([['общая'], ['площадь']])
    out1 = df1.copy(deep=True)
    out2 = df2.copy(deep=True)
    out1.columns = col1.append(col2).append(col3)
    out2.columns = col2.append(col3).append(col4)

    # Выводим несколько первых и последних строк основной таблицы
    ind = out1.index.tolist()
    print(out1.iloc[ind[:3] + ind[-2:]].to_string(index=False), '\n', '\n')
    # Выводим результаты анализа
    print(out2.to_string(index=False))
    return


def create_header(ws, n, row):
    ws.cell(row=row, column=3).value = 'участие'
    ws.cell(row=row + 2, column=3).value = 'да'
    ws.cell(row=row + 2, column=4).value = 'нет'
    ws.cell(row=row + 2, column=5).value = 'недейств'
    ws.cell(row=row, column=3).style = 'header_1'
    ws.cell(row=row + 2, column=3).style = 'small_head'
    ws.cell(row=row + 2, column=4).style = 'small_head'
    ws.cell(row=row + 2, column=5).style = 'small_head'
    ws.merge_cells(start_row=row, start_column=(3), end_row=row + 1, end_column=(5))
    for i in range(n):
        ws.cell(row=row, column=(6 + i * 3)).value = 'Вопрос ' + str(i + 1)
        ws.cell(row=row + 2, column=(6 + i * 3)).value = 'да'
        ws.cell(row=row + 2, column=(7 + i * 3)).value = 'возд'
        ws.cell(row=row + 2, column=(8 + i * 3)).value = 'нет'
        ws.cell(row=row, column=(6 + i * 3)).style = 'header_1'
        ws.cell(row=row + 2, column=(6 + i * 3)).style = 'small_head'
        ws.cell(row=row + 2, column=(7 + i * 3)).style = 'small_head'
        ws.cell(row=row + 2, column=(8 + i * 3)).style = 'small_head'
        ws.merge_cells(start_row=row, start_column=(6 + i * 3), end_row=row + 1, end_column=(8 + i * 3))
    ws.cell(row=row, column=(6 + n * 3)).value = 'голосов'
    ws.cell(row=row, column=(6 + n * 3)).style = 'header_1'
    ws.merge_cells(start_row=row, start_column=(6 + n * 3), end_row=row + 2, end_column=(6 + n * 3))
    column = ws.cell(row=row + 3, column=(6 + n * 3)).column_letter
    ws.column_dimensions[column].width = 14
    return


def safe_to_excel(df1, df2, dir_file, num_quest):

    # Создаем выходной файл
    template_file = str(Path(__file__).parents[1]) + r'/template/template.xlsx'
    output_file = filedialog.asksaveasfilename(
        defaultextension='.xlsx', filetypes=[("Excel files", '*.xlsx')],
        initialdir=dir_file,
        initialfile='Результаты голосования.xlsx',
        title="Выберите название выходного файла")
    # output_file = dir_file + '/Результаты голосования.xlsx'
    shutil.copy(template_file, output_file)

    # Настраиваем файл для записи
    book = load_workbook(output_file)
    writer = pd.ExcelWriter(output_file, engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    ws = writer.book.active

    # переставим столбцы в нужном порядке
    c = df1.columns.tolist()
    col = c[1:2] + c[:1] + c[3:] + c[2:3]
    df1 = df1[col]

    # записываем данные в файл
    result_row = df1.index.size+8
    df1.to_excel(writer, startrow=6, startcol=0, header=False, index=False)
    df2.to_excel(writer, startrow=result_row + 2, startcol=2, header=False, index=False)

    # форматирование таблицы
    create_header(ws, num_quest, 4)
    create_header(ws, num_quest, result_row)
    ws.cell(row=result_row, column=(6 + num_quest * 3)).value = 'общая\nплощадь'
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=(6 + num_quest * 3))
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=(6 + num_quest * 3))
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=(6 + num_quest * 3))
    ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')
    ws.cell(row=2, column=1).alignment = Alignment(horizontal='center')
    ws.cell(row=3, column=1).alignment = Alignment(horizontal='center')
    for row in range(7, df1.index.size+7):
        for col in range(1, 7 + num_quest * 3):
            ws.cell(row=row, column=col).style = 'value'
        ws.cell(row=row, column=2).style = 'name'
    for row in range(result_row + 3, result_row + 5):
        for col in range(3, 7 + num_quest * 3):
            ws.cell(row=row, column=col).style = 'value'

    # Задаем область печати
    print_area = ('A1:' + ws.cell(row=8, column=(6 + num_quest * 3)).column_letter
                  + str(result_row + 4))
    ws.print_area = print_area
    writer.save()
    return


def tszh_helper():
    member_list, directory = member_list_import()
    num_quest = number_of_q()
    i = 0
    while i == 0:
        ans = input('Ввести шаблон? (y/n)')
        if ans == 'y':
            print('задаем шаблон')
            output = template_enter(member_list, num_quest)
            i += 1
        elif ans == 'n':
            print('переходим к построчному вводу')
            output = single_line_enter(member_list, num_quest)
            i += 1
        else:
            print('некорректный ввод')

    # удаляем вспомогательный столбец - флаг обработки строки
    # и заменяем нули на пустые ячейки
    output = output.drop(columns='filled')
    mask = (output != 0)
    output = output.where(mask, None)

    analyse = result_analyse(output)
    print_result(output, analyse)
    safe_to_excel(output, analyse, directory, num_quest)
    print('\n', 'Обработка завершена')


if __name__ == '__main__':
    tszh_helper()
